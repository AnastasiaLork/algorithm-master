from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

def hour(ws, row_job, hours):
    for rowes in ws.rows:
        for cell in rowes:
            for row in range(row_job, row_job+1):
                for col in range(hours, hours+1):
                    char = get_column_letter(col)
                    if isinstance(ws[char+str(row)], MergedCell):
                        for x in ws.merged_cells.ranges:
                            if char + str(row) in x:
                                houries = str(x.start_cell.value)
                                return houries
                    elif ws[char+str(row)].value and not isinstance(ws[char+str(row)], MergedCell):
                        houries = ws[char+str(row)].value
                        return houries
